from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "커피24 견적서 AX TF.xlsx"
# Coffee service resources live under public/api/coffee/ so additional
# services (정수기, 스낵 등) can drop alongside without collision.
OUTPUT_DIR = ROOT / "public" / "api" / "coffee"

PURCHASE_OVERRIDES = {
    "카페스타 720": {
        "purchasePrice": 570000,
        "purchaseLabel": "카페스타 720\n수동세척 (주 2회 必)",
    },
    "TE-201U": {
        "purchasePrice": 600000,
        "purchaseLabel": "TE-201U\n수동세척 (주 2회 必)",
    },
    "HB-F09": {
        "purchasePrice": 1700000,
        "purchaseLabel": "하우스브란트 HB-F09\n자동세척(세척알약 1BOX증정)",
    },
    "HB-F12": {
        "purchasePrice": 2800000,
        "purchaseLabel": "칼렘 HB-F12\n자동세척(세척알약 1BOX증정)",
    },
    "E20T": {
        "purchasePrice": 1700000,
        "purchaseLabel": "칼렘 E20T\n자동세척 (세척알약 1BOX증정)",
    },
    "E30T": {
        "purchasePrice": 2000000,
        "purchaseLabel": "칼렘 E30T\n자동세척 (세척알약 1BOX증정)",
    },
    "E60T": {
        "purchasePrice": 2800000,
        "purchaseLabel": "칼렘 E60T\n자동세척(세척알약 1BOX증정)",
    },
    "JL29": {
        "purchasePrice": 2300000,
        "purchaseLabel": "제티노 JL29\n자동세척(세척알약 1BOX증정)",
    },
    "JL30": {
        "purchasePrice": 2460000,
        "purchaseLabel": "제티노 JL30\n자동세척(세척알약 1BOX증정)",
    },
    "X4": {
        "purchasePrice": 5500000,
        "purchaseLabel": "유라 X4\n자동세척 (세척알약 1BOX증정)",
    },
}


def clean_text(value):
    if value is None:
        return None
    return str(value).strip().strip('"')


def iter_rows(sheet, width: int):
    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = list(row[:width])
        if any(value not in (None, "") for value in values):
            yield values


def write_json(name: str, payload):
    path = OUTPUT_DIR / name
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return path


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(SOURCE, data_only=True)

    machines = []
    for name, description, spec, product_name, item_name in iter_rows(wb["머신정보"], 5):
        overrides = PURCHASE_OVERRIDES.get(name, {})
        machines.append({
            "name": name,
            "description": clean_text(description),
            "spec": clean_text(spec),
            "productName": clean_text(product_name) or name,
            "rentalItemName": clean_text(item_name) or "머신 단독 렌탈(4개월 1회 케어)",
            "purchasePrice": overrides.get("purchasePrice"),
            "purchaseLabel": overrides.get("purchaseLabel") or clean_text(product_name) or name,
            # Reserved for future product photo upload (see Machine.imageUrl in src/types/domain.ts)
            "imageUrl": None,
        })
    write_json("machines.json", machines)

    machine_prices = []
    for machine, term, included, price in iter_rows(wb["머신단가"], 4):
        machine_prices.append({
            "machine": machine,
            "term": int(term),
            "beansIncluded": included,
            "price": int(price),
        })
    write_json("machine-prices.json", machine_prices)

    care_cycles = [
        {"cycle": cycle, "extra": int(extra)}
        for cycle, extra in iter_rows(wb["케어주기"], 2)
    ]
    write_json("care-cycles.json", care_cycles)

    beans = [
        {"name": name, "price": int(price), "brand": brand, "imageUrl": None}
        for name, price, brand in iter_rows(wb["원두정보"], 3)
    ]
    write_json("beans.json", beans)

    reps = [
        {"name": name, "phone": phone, "email": email}
        for name, phone, email in iter_rows(wb["영업담당자"], 3)
    ]
    write_json("sales-reps.json", reps)

    write_json("constants.json", {
        "quoteValidDays": 15,
        "vatRate": 0.1,
        "shippingFeeVatIncluded": 3500,
        "freeShippingThresholdVatIncluded": 50000,
        "purchaseInstallFee": 70000,
        "setupVisitFee": 60000,
        "plumbInstallFee": 50000,
        "plumbKitFee": 25000,
        "ownershipTransferMonths": 36,
    })

    print(f"wrote {len(machines)} machines · {len(machine_prices)} prices · {len(beans)} beans · {len(reps)} reps → public/api/")


if __name__ == "__main__":
    main()
